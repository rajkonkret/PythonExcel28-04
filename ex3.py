from pprint import pprint

import openpyxl

wb = openpyxl.load_workbook('./data/videogamesales.xlsx')
ws = wb.active
print(ws)

ws = wb['vgsales']

print("Total number of rows: " + str(ws.max_row))  # total number of rows: 16328
print("Total number of columns: " + str(ws.max_column))  # Total number of columns: 10

print("Value in cell A1 is:", ws['A1'].value)  # Value in cell A1 is: Rank

values = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column)]
print(values)
# ['Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher', 'NA_Sales', 'EU_Sales', 'JP_Sales']

data = [ws.cell(row=i, column=2).value for i in range(2, 12)]
print(data)
# ['Wii Sports',
#  'Super Mario Bros.',
#  'Mario Kart Wii',
#  'Wii Sports Resort',\
#  'Pokemon Red/Pokemon Blue',
#  'Tetris',
#  'New Super Mario Bros.',
#  'Wii Play',
#  'New Super Mario Bros. Wii',
#  'Duck Hunt']
pprint(data)

# ['Wii Sports',
#  'Super Mario Bros.',
#  'Mario Kart Wii',
#  'Wii Sports Resort',
#  'Pokemon Red/Pokemon Blue',
#  'Tetris',
#  'New Super Mario Bros.',
#  'Wii Play',
#  'New Super Mario Bros. Wii',
#  'Duck Hunt']

my_list = list()  # pusta lista

for value in ws.iter_rows(
        min_row=1, max_row=11, min_col=1, max_col=6,
        values_only=True):
    my_list.append(value)
print(my_list)

for ele1, ele2, ele3, ele4, ele5, ele6 in my_list:
    print(f"{ele1:<8}{ele2:>35}{ele3:^10}{ele4:<10}{ele5:<15}{ele6:<15}")

ws["K1"] = "Sum of Sales"

wb.save('vidogamesales2.xlsx')
wb.close()
