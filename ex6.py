import openpyxl

wb = openpyxl.load_workbook('vidogamesales2.xlsx')
ws = wb.active

ws = wb['vgsales']

# nowy wiersz
new_row = (1, 'The Legend of Zelda', 1986, 'Action', 'Nintendo', 3.74, 0.93, 1.69, 0.14, 6.51, 6.5)
ws.append(new_row)

wb.save('videogamesales2.xlsx')
wb.close()

# sprawdzenie czy zapisał

# odczytanie ostaniego wiersza
values = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]
print(values)
# [1, 'The Legend of Zelda', 1986, 'Action', 'Nintendo', 3.74, 0.93, 1.69, 0.14, 6.51, 6.5]

# usunięcie jednego ostatniego wiersza
ws.delete_rows(ws.max_row, 1)

wb.save('videogamesales2.xlsx')
wb.close()
