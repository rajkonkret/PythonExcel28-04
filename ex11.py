import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('videogamesales2.xlsx')
ws = wb.active

ws = wb['Video Games Sales Data']

ws['A1'].font = Font(color='FF0000', bold=True, size=12)  # red
ws['A2'].font = Font(color='0000FF')  # blue

ws['A1'].fill = PatternFill('lightVertical', start_color='38e3ff')

wb.save('videogamesales2.xlsx')
wb.close()
